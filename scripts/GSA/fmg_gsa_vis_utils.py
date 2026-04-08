import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

# Plot the sensitivity indices
def plot_global_sensitivity_indices(S, E, N, HS, w_freq, GnP_list, params_list, file_path):
    """Plot variation of sensitivity indices over decades of shifted frequency

    Args:
        S (str): Sensitivity index type.
        E (str): Modulus type.
        N (int): Number of samples.
        HS (str): Hard segment value.
        w_freq (array-like): Array of shifted frequencies at which to evaluate the model.
        GnP_list (list): List of GnP values.
        params_list (list): List of parameter names.
        file_path (dict): File path for saving plots.
    """
    # Load the sensitivity indices data frames
    S_Ep_dfs = {}
    for GnP_idx in range(len(GnP_list)):
        S_Ep_dfs[GnP_idx] = pd.read_excel(f'{file_path["save_path"]}/{S}_{E}_{HS}_{GnP_list[GnP_idx]}.xlsx', sheet_name=f'N={N}')
    
    # Create subplots
    _, axs = plt.subplots(3, 2, figsize=(10, 5), dpi=600)
    axs = axs.flatten()
    subplot_idx = [0, 2, 4, 1, 3, 5]

    # y axis names
    ylabels = [r"{E^{\prime}, E_{c_1}}", r"{E^{\prime}, \tau_{c_1}}", r"{E^{\prime}, \alpha_1}",
               r"{E^{\prime}, E_{c_2}}", r"{E^{\prime}, \tau_{c_2}}", r"{E^{\prime}, \alpha_2}"]
    
    for GnP_idx in range(len(GnP_list)):
        S_Ep_df = S_Ep_dfs[GnP_idx]
        for param_idx in range(len(params_list)):
            axs[subplot_idx[param_idx]].plot(w_freq, S_Ep_df[params_list[param_idx]], label=GnP_list[GnP_idx], linestyle='-')
            axs[subplot_idx[param_idx]].set_xscale('log')
            axs[subplot_idx[param_idx]].set_ylabel(fr'$S1_{ylabels[param_idx]}$')
            axs[subplot_idx[param_idx]].set_xlim([0.5*w_freq[0], 1.5*w_freq[-1]])
            axs[subplot_idx[param_idx]].set_xticks([1e-8, 1e-6, 1e-4, 1e-2, 1e0, 1e2])
            axs[subplot_idx[param_idx]].set_ylim([-0.05, 1.05])

    axs[-2].set_xlabel(r'$\omega a_{T} \ (rad/s)$')
    axs[-1].set_xlabel(r'$\omega a_{T} \ (rad/s)$')
    axs[0].legend(ncols=2)

    plt.tight_layout()
    plt.savefig(f'{file_path["save_path"]}/{S}_{E}_{HS}.png', dpi=600)
    plt.show()

# Define a function to read a specified range of cells from an Excel file and return it as a NumPy array
def read_excel_range(filename, cell_range):
    """Read a specified range of cells from an Excel file and return it as a NumPy array.

    Args:
        filename (str): Path to the Excel file.
        cell_range (str): Range of cells to read.

    Returns:
        np.ndarray: A NumPy array containing the values from the specified cell range.
    """
    wb = load_workbook(filename=filename, data_only=True, read_only=True)
    ws = wb["Sheet1"]

    min_col, min_row, max_col, max_row = range_boundaries(cell_range)

    data = []
    for r in range(min_row, max_row + 1):
        row_vals = []
        for c in range(min_col, max_col + 1):
            v = ws.cell(row=r, column=c).value
            row_vals.append(np.nan if v is None else float(v))
        data.append(row_vals)

    wb.close()
    return np.array(data, dtype=float)

# Define a function to plot the L-infinity norm of the sensitivity indices for each parameter, grouped by GnP content
def plot_Linf_grouped(mean_S,
                      std_S,
                      ylabel,
                      params_list=[r"$E_{c_1}$", r"$\tau_{c_1}$", r"$\alpha_1$", r"$\beta_1$", r"$E_{c_2}$", r"$\tau_{c_2}$", r"$\alpha_2$"],
                      legend_labels=("20% HSWF", "30% HSWF", "40% HSWF"),
                      figsize=(6, 4),
                      save_path=None):
    """Plot the L-infinity norm of the sensitivity indices for each parameter, grouped by GnP content.

    Args:
        mean_S (numpy.ndarray): Array of mean sensitivity indices.
        std_S (numpy.ndarray): Array of standard deviations of sensitivity indices.
        ylabel (str): Label for the y-axis.
        params_list (list): List of parameter names.
        legend_labels (tuple): Labels for the legend.
        figsize (tuple): Figure size.
        save_path (str, optional): Path to save the plot. Defaults to None.
    """
    y = mean_S.T
    err = std_S.T

    sort_order = np.argsort(np.mean(y, axis=1))[::-1]
    sortedY = y[sort_order, :]
    sortedErr = err[sort_order, :]
    sortedLabels = [params_list[i] for i in sort_order]

    ngroups, nbars = sortedY.shape

    _ = plt.figure(figsize=figsize)
    ax = plt.gca()

    group_x = np.arange(ngroups)
    total_group_width = 0.8
    bar_w = total_group_width / nbars
    offsets = (np.arange(nbars) - (nbars - 1) / 2.0) * bar_w

    bars = []
    for i in range(nbars):
        x_i = group_x + offsets[i]
        b = ax.bar(x_i, sortedY[:, i], width=bar_w, label=legend_labels[i])
        bars.append(b)

        ax.errorbar(
            x_i,
            sortedY[:, i],
            yerr=sortedErr[:, i],
            fmt="none",
            ecolor="k",
            capsize=0
        )

    ax.set_ylabel(ylabel)
    ax.set_xlabel(r"Model Parameters")
    ax.set_ylim(0.0, 1.05)

    ax.set_xticks(group_x)
    ax.set_xticklabels(sortedLabels)
    ax.legend(frameon=False)

    plt.tight_layout()
    plt.savefig(f"{save_path}.png", dpi=300)
    plt.show()