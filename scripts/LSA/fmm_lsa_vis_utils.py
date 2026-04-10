import numpy as np
import matplotlib.pyplot as plt

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

def plot_local_sensitivity_indices(E_type, HS, GnP_list, file_path):
    """Plots the local sensitivity indices for the specified modulus type.
    Args:
        E_type (str): Type of modulus ('Ep', 'Epp', or 'Ecomplex').
        HS (int): Heat treatment condition.
        GnP_list (list): List of GnP loading conditions.
        file_path (dict): Dictionary containing paths to save the plot and load data."""
    fig, axs = plt.subplots(4, 2, figsize=(10, 6.67), dpi=600)
    axs = axs.flatten()
    subplot_idx = [0, 2, 4, 6, 1, 3, 5]

    sym_map = {'Ep': r'E^{\prime}', 'Epp': r'E^{\prime\prime}', 'Ecomplex': r'E^*'}
    mod_sym = sym_map.get(E_type, r'E')

    param_symbols = [r'E_{c_1}', r'\tau_{c_1}', r'\alpha_1', r'\beta_1',
                    r'E_{c_2}', r'\tau_{c_2}', r'\alpha_2']

    ylabels = [fr'$\bar{{S}}_{{{mod_sym}, {p}}}$' for p in param_symbols]

    for GnP in GnP_list:
        file_name = f'{file_path["save_path"]}/{HS}HS_{GnP}.npz'
        array_key = f'all_lsi_{E_type}_{HS}HS_{GnP}'
        data = np.load(file_name)
        mean_std_indices = data[array_key]
        w_freq = mean_std_indices[-1, :]

        for i in range(len(param_symbols)):
            ax = axs[subplot_idx[i]]
            
            mean_val = mean_std_indices[i * 2, :]
            
            ax.plot(w_freq, mean_val, label=f'{GnP}', linestyle='-', linewidth=1.5)
            
            ax.set_xscale('log')
            ax.set_ylabel(ylabels[i])
            ax.set_xlim([0.5 * w_freq[0], 1.5 * w_freq[-1]])
            ax.set_xticks([1e-8, 1e-6, 1e-4, 1e-2, 1e0, 1e2])

    axs[-2].set_xlabel(r'$\omega a_{T} \ (rad/s)$')
    axs[-3].set_xlabel(r'$\omega a_{T} \ (rad/s)$')
    axs[0].legend(GnP_list, loc='best')
    fig.delaxes(axs[-1])
    axs[0].legend()
    plt.tight_layout()

    save_file = f'{file_path["save_path"]}/LSI_{E_type}_{HS}HS.png'
    plt.savefig(save_file, dpi=600, bbox_inches='tight')
    plt.show()

# Define a function to read a specified range of cells from an Excel file and return it as a NumPy array
def read_excel_range(filename, cell_range):
    """Read a specified range of cells from an Excel file and return it as a NumPy array.

    Args:
        filename (str): Path to the Excel file.
        cell_range (str): Range of cells to read.

    Returns:
        np.ndarray: A NumPy array containing the values from the specified cell range.
    """
    wb = load_workbook(filename=filename, data_only=True, read_only=True)
    ws = wb["Sheet1"]

    min_col, min_row, max_col, max_row = range_boundaries(cell_range)

    data = []
    for r in range(min_row, max_row + 1):
        row_vals = []
        for c in range(min_col, max_col + 1):
            v = ws.cell(row=r, column=c).value
            row_vals.append(np.nan if v is None else float(v))
        data.append(row_vals)

    wb.close()
    return np.array(data, dtype=float)

# Define a function to plot the L-infinity norm of the sensitivity indices for each parameter, grouped by GnP content
def plot_L1_grouped(mean_S,
                    std_S,
                    ylabel,
                    params_list=[r"$E_{c_1}$", r"$\tau_{c_1}$", r"$\alpha_1$", r"$\beta_1$", r"$E_{c_2}$", r"$\tau_{c_2}$", r"$\alpha_2$"],
                    legend_labels=("20% HSWF", "30% HSWF", "40% HSWF"),
                    figsize=(6, 4),
                    save_path=None):
    """Plot the L-infinity norm of the sensitivity indices for each parameter, grouped by GnP content.

    Args:
        mean_S (numpy.ndarray): Array of mean sensitivity indices.
        std_S (numpy.ndarray): Array of standard deviations of sensitivity indices.
        ylabel (str): Label for the y-axis.
        params_list (list): List of parameter names.
        legend_labels (tuple): Labels for the legend.
        figsize (tuple): Figure size.
        save_path (str, optional): Path to save the plot. Defaults to None.
    """
    y = mean_S.T
    err = std_S.T

    sort_order = np.argsort(np.mean(y, axis=1))[::-1]
    sortedY = y[sort_order, :]
    sortedErr = err[sort_order, :]
    sortedLabels = [params_list[i] for i in sort_order]

    ngroups, nbars = sortedY.shape

    _ = plt.figure(figsize=figsize)
    ax = plt.gca()

    group_x = np.arange(ngroups)
    total_group_width = 0.8
    bar_w = total_group_width / nbars
    offsets = (np.arange(nbars) - (nbars - 1) / 2.0) * bar_w

    bars = []
    for i in range(nbars):
        x_i = group_x + offsets[i]
        b = ax.bar(x_i, sortedY[:, i], width=bar_w, label=legend_labels[i])
        bars.append(b)

        ax.errorbar(
            x_i,
            sortedY[:, i],
            yerr=sortedErr[:, i],
            fmt="none",
            ecolor="k",
            capsize=0
        )

    ax.set_ylabel(ylabel)
    ax.set_xlabel(r"Model Parameters")

    ax.set_xticks(group_x)
    ax.set_xticklabels(sortedLabels)
    ax.legend(frameon=False)

    plt.tight_layout()
    plt.savefig(f"{save_path}.png", dpi=300)
    plt.show()